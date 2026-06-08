import re
import cv2
import numpy as np
from pathlib import Path
import easyocr
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
# ДИАГНОСТИКА
# ─────────────────────────────────────────────────────────────────────────────

def _save_debug_stages(img: np.ndarray, binary: np.ndarray,
                        h_lines: np.ndarray, v_lines: np.ndarray,
                        row_ys: list[int], col_xs: list[int],
                        out_dir: Path) -> None:
    """Сохраняет промежуточные изображения каждого этапа обработки."""
    h, w = img.shape[:2]

    # 1. Бинаризованное
    cv2.imwrite(str(out_dir / "debug_1_binary.png"), binary)

    # 2. Выделенные горизонтальные линии
    cv2.imwrite(str(out_dir / "debug_2_hlines.png"), h_lines)

    # 3. Выделенные вертикальные линии
    cv2.imwrite(str(out_dir / "debug_3_vlines.png"), v_lines)

    # 4. Горизонтальная проекция (h_proj) как график
    h_proj = h_lines.sum(axis=1).astype(float)
    h_proj_img = np.ones((h, 300), dtype=np.uint8) * 255
    if h_proj.max() > 0:
        for y, val in enumerate(h_proj):
            bar_len = int(val / h_proj.max() * 290)
            cv2.line(h_proj_img, (0, y), (bar_len, y), 0, 1)
    cv2.imwrite(str(out_dir / "debug_4_hproj.png"), h_proj_img)

    # 5. Вертикальная проекция (v_proj) как график
    v_proj = v_lines.sum(axis=0).astype(float)
    v_proj_img = np.ones((300, w), dtype=np.uint8) * 255
    if v_proj.max() > 0:
        for x, val in enumerate(v_proj):
            bar_len = int(val / v_proj.max() * 290)
            cv2.line(v_proj_img, (x, 300), (x, 300 - bar_len), 0, 1)
    cv2.imwrite(str(out_dir / "debug_5_vproj.png"), v_proj_img)

    # 6. Итоговая сетка поверх оригинала
    grid_img = img.copy()
    for y in row_ys:
        cv2.line(grid_img, (0, y), (w, y), (0, 0, 255), 2)
    for x in col_xs:
        cv2.line(grid_img, (x, 0), (x, h), (0, 255, 0), 2)
    cv2.imwrite(str(out_dir / "debug_6_grid.png"), grid_img)

    print(f"\n   [debug] Сохранено в {out_dir}:")
    print(f"           debug_1_binary.png  — бинаризация")
    print(f"           debug_2_hlines.png  — горизонтальные линии")
    print(f"           debug_3_vlines.png  — вертикальные линии")
    print(f"           debug_4_hproj.png   — проекция по Y")
    print(f"           debug_5_vproj.png   — проекция по X")
    print(f"           debug_6_grid.png    — итоговая сетка\n")


# ─────────────────────────────────────────────────────────────────────────────
# 1. Склейка
# ─────────────────────────────────────────────────────────────────────────────

def _merge_images(img_paths: list[Path]) -> np.ndarray:
    """Склеивает PNG вертикально, приводя к одинаковой ширине."""
    images = [cv2.imread(str(p)) for p in img_paths]
    images = [img for img in images if img is not None]
    if not images:
        raise ValueError("Не удалось загрузить ни одного изображения.")

    max_w = max(img.shape[1] for img in images)
    resized = []
    for img in images:
        h, w = img.shape[:2]
        if w != max_w:
            img = cv2.resize(
                img, (max_w, int(h * max_w / w)),
                interpolation=cv2.INTER_AREA,
            )
        resized.append(img)

    return np.vstack(resized)


# ─────────────────────────────────────────────────────────────────────────────
# 2. Детекция сетки
# ─────────────────────────────────────────────────────────────────────────────

def _detect_table_grid(
    img: np.ndarray,
    debug_dir: Path | None = None,
) -> tuple[list[int], list[int]]:
    """
    Детектирует линии таблицы.

    Стратегия:
      - Адаптивная бинаризация (устойчива к фону любого цвета)
      - Отдельные ядра для H и V линий с размерами,
        не зависящими от полной высоты склеенного изображения
      - После проекции — подавление ложных пиков через медианный фильтр
    """
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    # Адаптивная бинаризация: локальный порог в окне 15×15
    binary = cv2.adaptiveThreshold(
        gray, 255,
        cv2.ADAPTIVE_THRESH_MEAN_C,
        cv2.THRESH_BINARY_INV,
        15, 4,
    )

    h, w = binary.shape

    # ── Горизонтальные линии ──────────────────────────────────────────────
    # Длина ядра = 40% ширины (достаточно для сплошной линии через всю таблицу)
    h_len = max(int(w * 0.4), 30)
    hk    = cv2.getStructuringElement(cv2.MORPH_RECT, (h_len, 1))
    h_lines = cv2.morphologyEx(binary, cv2.MORPH_OPEN, hk, iterations=1)

    # ── Вертикальные линии ────────────────────────────────────────────────
    # Фиксированная высота ядра = 30 px — не зависит от высоты склейки
    # Это позволяет находить вертикальные сегменты внутри каждого фрагмента
    v_len = 30
    vk    = cv2.getStructuringElement(cv2.MORPH_RECT, (1, v_len))
    v_lines = cv2.morphologyEx(binary, cv2.MORPH_OPEN, vk, iterations=1)

    # ── Проекция → пики ───────────────────────────────────────────────────
    def _find_peaks(proj: np.ndarray, min_gap: int, rel_thresh: float) -> list[int]:
        """Находит позиции пиков в проекции."""
        if proj.max() == 0:
            return []
        thresh    = proj.max() * rel_thresh
        positions = np.where(proj > thresh)[0].tolist()
        if not positions:
            return []
        groups, cur = [], [positions[0]]
        for p in positions[1:]:
            if p - cur[-1] <= min_gap:
                cur.append(p)
            else:
                groups.append(cur)
                cur = [p]
        groups.append(cur)
        return [int(np.mean(g)) for g in groups]

    h_proj = h_lines.sum(axis=1).astype(float)
    v_proj = v_lines.sum(axis=0).astype(float)

    row_ys = _find_peaks(h_proj, min_gap=6,  rel_thresh=0.25)
    col_xs = _find_peaks(v_proj, min_gap=6,  rel_thresh=0.10)

    # ── Фильтрация: убираем линии слишком близко к краям ─────────────────
    margin = 5
    row_ys = [y for y in row_ys if margin < y < h - margin]
    col_xs = [x for x in col_xs if margin < x < w - margin]

    # ── Debug ─────────────────────────────────────────────────────────────
    if debug_dir is not None:
        _save_debug_stages(img, binary, h_lines, v_lines, row_ys, col_xs, debug_dir)

    return row_ys, col_xs


# ─────────────────────────────────────────────────────────────────────────────
# 3. Ячейки
# ─────────────────────────────────────────────────────────────────────────────

def _build_cells(
    row_ys: list[int],
    col_xs: list[int],
) -> list[list[tuple[int, int, int, int]]]:
    """Строит сетку ячеек (x0, y0, x1, y1) из линий."""
    cells = []
    for i in range(len(row_ys) - 1):
        row = []
        for j in range(len(col_xs) - 1):
            row.append((col_xs[j], row_ys[i], col_xs[j + 1], row_ys[i + 1]))
        cells.append(row)
    return cells


# ─────────────────────────────────────────────────────────────────────────────
# 4. OCR ячейки
# ─────────────────────────────────────────────────────────────────────────────

def _ocr_cell(
    img: np.ndarray,
    x0: int, y0: int,
    x1: int, y1: int,
    reader: easyocr.Reader,
    padding: int = 4,
) -> str:
    """
    Вырезает ячейку и распознаёт текст.
    Если обычный проход пустой — пробует повёрнутое изображение
    (для вертикального текста в шапке).
    """
    H, W = img.shape[:2]
    cx0, cy0 = max(x0 + padding, 0), max(y0 + padding, 0)
    cx1, cy1 = min(x1 - padding, W), min(y1 - padding, H)

    if cx1 <= cx0 or cy1 <= cy0:
        return ""

    cell_img = img[cy0:cy1, cx0:cx1]
    if cell_img.shape[0] < 5 or cell_img.shape[1] < 5:
        return ""

    # Увеличиваем для лучшего OCR
    scale    = 2
    cell_img = cv2.resize(
        cell_img,
        (cell_img.shape[1] * scale, cell_img.shape[0] * scale),
        interpolation=cv2.INTER_CUBIC,
    )

    def _read(im: np.ndarray) -> str:
        res = reader.readtext(im, detail=0, paragraph=True)
        return " ".join(r.strip() for r in res if r.strip())

    text = _read(cell_img)

    # Попытка с поворотом — для вертикального текста шапки
    if not text:
        text = _read(cv2.rotate(cell_img, cv2.ROTATE_90_COUNTERCLOCKWISE))
    if not text:
        text = _read(cv2.rotate(cell_img, cv2.ROTATE_90_CLOCKWISE))

    return text


# ─────────────────────────────────────────────────────────────────────────────
# 5. Основная функция
# ─────────────────────────────────────────────────────────────────────────────

def convert_to_xlsx(pdfname: str) -> Path:
    """
    Склеивает PNG → детектирует сетку → OCR каждой ячейки → xlsx без дизайна.
    """
    # ── Пути ─────────────────────────────────────────────────────────────
    folder_name = pdfname.split(".")[0]
    base_dir    = Path("../../tables") / folder_name
    input_dir   = base_dir / "tables"
    output_file = base_dir / f"{folder_name}_tables.xlsx"
    debug_dir   = base_dir / "debug"
    debug_dir.mkdir(parents=True, exist_ok=True)

    if not input_dir.exists():
        raise FileNotFoundError(f"Папка не найдена: {input_dir.resolve()}")

    # ── Файлы ─────────────────────────────────────────────────────────────
    file_pattern = re.compile(r"^table_.*_(\d+)\.png$")
    matched = sorted(
        [
            (int(m.group(1)), f)
            for f in input_dir.iterdir()
            if (m := file_pattern.match(f.name))
        ],
        key=lambda t: t[0],
    )
    if not matched:
        raise ValueError(f"Файлы не найдены в {input_dir.resolve()}")

    img_paths = [p for _, p in matched]

    # ── Склейка ───────────────────────────────────────────────────────────
    print("🔗 Склейка изображений ...", flush=True)
    merged = _merge_images(img_paths)
    cv2.imwrite(str(debug_dir / "debug_0_merged.png"), merged)
    print(f"   Размер: {merged.shape[1]}×{merged.shape[0]} px")

    # ── Сетка ─────────────────────────────────────────────────────────────
    print("📐 Детекция сетки ...", flush=True)
    row_ys, col_xs = _detect_table_grid(merged, debug_dir=debug_dir)

    n_rows = len(row_ys) - 1
    n_cols = len(col_xs) - 1
    print(f"   Строк: {n_rows}  |  Столбцов: {n_cols}")

    if n_rows < 1 or n_cols < 1:
        raise ValueError(
            f"Сетка не построена (строк={n_rows}, столбцов={n_cols}).\n"
            f"Проверьте debug-изображения в {debug_dir.resolve()}"
        )

    cells = _build_cells(row_ys, col_xs)

    # ── OCR ───────────────────────────────────────────────────────────────
    print("🔍 OCR ...", flush=True)
    reader     = easyocr.Reader(["ru", "en"], gpu=False, verbose=False)
    table_data = []

    for r_idx, row_cells in enumerate(cells):
        row_texts = [
            _ocr_cell(merged, x0, y0, x1, y1, reader)
            for x0, y0, x1, y1 in row_cells
        ]
        table_data.append(row_texts)
        print(f"   [{r_idx+1}/{n_rows}] {row_texts}")

    # ── Excel (без дизайна) ───────────────────────────────────────────────
    print("💾 Запись в Excel ...", flush=True)
    wb = Workbook()
    ws = wb.active
    ws.title = folder_name[:31]

    for r_idx, row in enumerate(table_data, start=1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    for c_idx in range(1, n_cols + 1):
        max_len = max(
            (len(str(table_data[r][c_idx-1]))
             for r in range(len(table_data))
             if c_idx-1 < len(table_data[r])),
            default=8,
        )
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 4, 60)

    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)

    print(f"\n✅ {output_file.resolve()}")
    print(f"   Строк: {len(table_data)}  |  Столбцов: {n_cols}")