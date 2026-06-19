from img2table.ocr import EasyOCR
from img2table.document import PDF
import os
import openpyxl
from openpyxl.utils import get_column_letter

_ocr = None

def get_ocr():
    global _ocr
    if _ocr is None:
        _ocr = EasyOCR(lang=["en", "ru"])
    return _ocr

# Импортируйте ваши модули для PDF и mkdir, если они не импортированы ранее:
# from ВашаБиблиотека import PDF, mkdir

def extract_data(pdfname):
    base_name = pdfname.split('.')[0]

    dir_path = f'../../tables/{base_name}'

    temp_output_path = f'{dir_path}/{base_name}_temp.xlsx'
    final_output_path = f'{dir_path}/{base_name}.xlsx'

    pdf = PDF(src=pdfname)
    pdf.to_xlsx(temp_output_path,
                ocr=get_ocr(),
                implicit_columns=True,
                borderless_tables=False,
                min_confidence=70)

    wb_source = openpyxl.load_workbook(temp_output_path)
    wb_final = openpyxl.Workbook()
    ws_dest = wb_final.active
    ws_dest.title = "Merged_Sheets"

    # Определяем эталонную ширину (количество колонок) по первому листу
    if wb_source.worksheets:
        base_col_count = wb_source.worksheets[0].max_column
    else:
        base_col_count = 1

    current_start_row = 1

    # Задаем ширину столбцов по первому листу
    first_sheet = wb_source.worksheets[0]
    for c in range(1, base_col_count + 1):
        col_letter = get_column_letter(c)
        if first_sheet.column_dimensions[col_letter].width:
            ws_dest.column_dimensions[col_letter].width = first_sheet.column_dimensions[col_letter].width

    for ws_source in wb_source.worksheets:
        max_row = ws_source.max_row
        source_col_count = ws_source.max_column

        # Вычисляем, сколько ПУСТЫХ столбцов нужно добавить СЛЕВА.
        # Если текущая таблица шире первой, отступ равен 0.
        left_padding = max(0, base_col_count - source_col_count)

        # Итоговая ширина строки в финальном листе
        current_max_col = max(source_col_count, base_col_count)

        # Если текущая таблица шире первой, расширяем ширину новых колонок в итоговом листе
        for c in range(base_col_count + 1, source_col_count + 1):
            col_letter = get_column_letter(c)
            if ws_source.column_dimensions[col_letter].width:
                ws_dest.column_dimensions[col_letter].width = ws_source.column_dimensions[col_letter].width

        # Копируем данные, стили и высоту строк
        for r in range(1, max_row + 1):
            dest_row_idx = current_start_row + r - 1

            if ws_source.row_dimensions[r].height:
                ws_dest.row_dimensions[dest_row_idx].height = ws_source.row_dimensions[r].height

            # Берем стиль границ из первой доступной ячейки текущей строки для пустых мест
            first_valid_border = None
            if source_col_count >= 1:
                first_valid_border = ws_source.cell(row=r, column=1).border

            # Заполняем строку слева направо до целевого максимума
            for c in range(1, current_max_col + 1):
                dest_cell = ws_dest.cell(row=dest_row_idx, column=c)

                if c <= left_padding:
                    # 1. ЗОНА ОТСТУПА: Добавляем пустые столбцы СЛЕВА
                    dest_cell.value = ""  # Явно инициализируем ячейку
                    if first_valid_border:
                        dest_cell.border = openpyxl.styles.Border(
                            **{k: v for k, v in first_valid_border.__dict__.items() if not k.startswith('_')})
                else:
                    # 2. ЗОНА ДАННЫХ: Копируем данные с текущего листа (со смещением на left_padding)
                    source_col_idx = c - left_padding
                    source_cell = ws_source.cell(row=r, column=source_col_idx)
                    dest_cell.value = source_cell.value

                    if source_cell.has_style:
                        dest_cell.font = openpyxl.styles.Font(
                            **{k: v for k, v in source_cell.font.__dict__.items() if not k.startswith('_')})
                        dest_cell.border = openpyxl.styles.Border(
                            **{k: v for k, v in source_cell.border.__dict__.items() if not k.startswith('_')})
                        dest_cell.fill = openpyxl.styles.PatternFill(
                            **{k: v for k, v in source_cell.fill.__dict__.items() if not k.startswith('_')})
                        dest_cell.alignment = openpyxl.styles.Alignment(
                            **{k: v for k, v in source_cell.alignment.__dict__.items() if not k.startswith('_')})

        current_start_row += max_row

    wb_source.close()
    wb_final.save(final_output_path)

    if os.path.exists(temp_output_path):
        os.remove(temp_output_path)

    print(f"Таблицы склеены. Пустые столбцы добавлены слева: {final_output_path}")
    return final_output_path
